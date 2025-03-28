import openpyxl as op
import random

# Exercício 6


wb = op.Workbook()
ws = wb.active
ws.title = 'Vendas'
ws.append(['Produto', 'Quantidade', 'Preço'])
produtos = ['Linguiça', 'Desodorante', 'Shampoo', 'Queijo', 'Manteiga']
for linha in range(2, 7):
    ws.cell(linha, 1, produtos[linha - 2])
    ws.cell(linha, 2, value=random.randint(1,20))
    ws.cell(linha, 3, value=random.randint(10,17))

wb.save('atividade_avaliativa_31-01/ex3/dados.xlsx')


# Exercício 7


wb = op.load_workbook('atividade_avaliativa_31-01/ex3/dados.xlsx')
ws = wb.active

for row in ws.iter_rows(values_only=True):
    print(row)


# Exercício 8

aux_a = op.Workbook()
aux = aux_a.active
aux.title = 'Vendas Filtradas'

wb = op.load_workbook('atividade_avaliativa_01-31/ex3/dados.xlsx')
ws = wb.active
line = 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if int(row[1]) * int(row[2]) > 100:
        aux.cell(line, 1, value=row[0])
        aux.cell(line, 2, value=row[1])
        aux.cell(line, 3, value=row[2])
        line += 1

aux_a.save('atividade_avaliativa_01-31/ex3/vendas_filtradas.xlsx')